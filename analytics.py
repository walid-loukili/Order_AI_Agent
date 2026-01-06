"""
Analytics Module
Advanced statistics, predictions, fraud detection, and reporting.
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
import io
import base64

# Fix Windows console encoding
sys.stdout.reconfigure(encoding='utf-8')

from database import DatabaseManager
from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()


class ClientHistory:
    """Manage client order history for smart reordering."""
    
    def __init__(self, db: DatabaseManager):
        self.db = db
    
    def get_client_preferences(self, client_name):
        """Get client's order history and preferences."""
        cursor = self.db.connection.cursor()
        
        # Get all orders for this client
        cursor.execute("""
            SELECT c.*, p.type as produit_type
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN produits p ON c.produit_id = p.id
            WHERE cl.nom LIKE ?
            ORDER BY c.created_at DESC
        """, (f"%{client_name}%",))
        
        orders = [dict(row) for row in cursor.fetchall()]
        
        if not orders:
            return None
        
        # Analyze preferences
        preferences = {
            "client_name": client_name,
            "total_orders": len(orders),
            "first_order": orders[-1]["created_at"] if orders else None,
            "last_order": orders[0]["created_at"] if orders else None,
            "favorite_products": self._get_favorite_products(orders),
            "average_quantity": self._get_average_quantity(orders),
            "typical_order": self._get_typical_order(orders),
            "order_frequency_days": self._calculate_order_frequency(orders),
            "total_spent": sum(o.get("prix_total") or 0 for o in orders)
        }
        
        return preferences
    
    def _get_favorite_products(self, orders):
        """Get most ordered products."""
        product_counts = defaultdict(int)
        for order in orders:
            if order.get("produit_type"):
                product_counts[order["produit_type"]] += 1
        return dict(sorted(product_counts.items(), key=lambda x: x[1], reverse=True))
    
    def _get_average_quantity(self, orders):
        """Calculate average order quantity."""
        quantities = [o.get("quantite") or 0 for o in orders if o.get("quantite")]
        return sum(quantities) / len(quantities) if quantities else 0
    
    def _get_typical_order(self, orders):
        """Get the most common order configuration."""
        if not orders:
            return None
        
        # Return the most recent validated order as typical
        for order in orders:
            if order.get("statut") == "validee":
                return {
                    "produit_type": order.get("produit_type"),
                    "nature_produit": order.get("nature_produit"),
                    "quantite": order.get("quantite"),
                    "unite": order.get("unite")
                }
        return None
    
    def _calculate_order_frequency(self, orders):
        """Calculate average days between orders."""
        if len(orders) < 2:
            return None
        
        dates = []
        for order in orders:
            try:
                if order.get("created_at"):
                    dates.append(datetime.fromisoformat(order["created_at"].replace("Z", "")))
            except:
                pass
        
        if len(dates) < 2:
            return None
        
        dates.sort(reverse=True)
        gaps = [(dates[i] - dates[i+1]).days for i in range(len(dates)-1)]
        return sum(gaps) / len(gaps) if gaps else None
    
    def suggest_reorder(self, client_name, partial_order_data):
        """Suggest order details based on client history."""
        preferences = self.get_client_preferences(client_name)
        
        if not preferences or not preferences.get("typical_order"):
            return partial_order_data
        
        typical = preferences["typical_order"]
        
        # Fill in missing fields from history
        if not partial_order_data.get("type_produit") and typical.get("produit_type"):
            partial_order_data["type_produit"] = typical["produit_type"]
        
        if not partial_order_data.get("nature_produit") and typical.get("nature_produit"):
            partial_order_data["nature_produit"] = typical["nature_produit"]
        
        if not partial_order_data.get("quantite") and typical.get("quantite"):
            partial_order_data["quantite"] = typical["quantite"]
        
        if not partial_order_data.get("unite") and typical.get("unite"):
            partial_order_data["unite"] = typical["unite"]
        
        partial_order_data["suggestion_source"] = "client_history"
        
        return partial_order_data


class Analytics:
    """Advanced analytics and statistics."""
    
    def __init__(self, db: DatabaseManager):
        self.db = db
    
    def get_dashboard_stats(self):
        """Get comprehensive dashboard statistics."""
        cursor = self.db.connection.cursor()
        
        stats = {}
        
        # Basic counts
        cursor.execute("SELECT COUNT(*) FROM commandes")
        stats["total_orders"] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM commandes WHERE statut = 'en_attente'")
        stats["pending_orders"] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM commandes WHERE statut = 'validee'")
        stats["validated_orders"] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM commandes WHERE statut = 'rejetee'")
        stats["rejected_orders"] = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM clients")
        stats["total_clients"] = cursor.fetchone()[0]
        
        # Revenue stats
        cursor.execute("SELECT SUM(prix_total), AVG(prix_total) FROM commandes WHERE prix_total IS NOT NULL")
        result = cursor.fetchone()
        stats["total_revenue"] = result[0] or 0
        stats["avg_order_value"] = result[1] or 0
        
        # Quantity stats
        cursor.execute("SELECT SUM(quantite), AVG(quantite) FROM commandes WHERE quantite IS NOT NULL")
        result = cursor.fetchone()
        stats["total_quantity"] = result[0] or 0
        stats["avg_quantity"] = result[1] or 0
        
        # Orders by product type
        cursor.execute("""
            SELECT p.type, COUNT(*) as count, SUM(c.quantite) as total_qty
            FROM commandes c
            LEFT JOIN produits p ON c.produit_id = p.id
            WHERE p.type IS NOT NULL
            GROUP BY p.type
            ORDER BY count DESC
        """)
        stats["orders_by_product"] = [dict(row) for row in cursor.fetchall()]
        
        # Orders by status
        cursor.execute("""
            SELECT statut, COUNT(*) as count
            FROM commandes
            GROUP BY statut
        """)
        stats["orders_by_status"] = {row["statut"]: row["count"] for row in cursor.fetchall()}
        
        # Top clients
        cursor.execute("""
            SELECT cl.nom, COUNT(*) as order_count, SUM(c.prix_total) as total_spent
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            GROUP BY cl.id
            ORDER BY order_count DESC
            LIMIT 10
        """)
        stats["top_clients"] = [dict(row) for row in cursor.fetchall()]
        
        # Recent trends (last 30 days)
        cursor.execute("""
            SELECT DATE(created_at) as date, COUNT(*) as count
            FROM commandes
            WHERE created_at >= datetime('now', '-30 days')
            GROUP BY DATE(created_at)
            ORDER BY date
        """)
        stats["daily_orders"] = [dict(row) for row in cursor.fetchall()]
        
        # Monthly trends
        cursor.execute("""
            SELECT strftime('%Y-%m', created_at) as month, 
                   COUNT(*) as count,
                   SUM(prix_total) as revenue
            FROM commandes
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month DESC
            LIMIT 12
        """)
        stats["monthly_trends"] = [dict(row) for row in cursor.fetchall()]
        
        return stats
    
    def get_orders_dataframe(self):
        """Get orders as pandas DataFrame for analysis."""
        cursor = self.db.connection.cursor()
        cursor.execute("""
            SELECT c.*, cl.nom as client_nom, p.type as produit_type
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN produits p ON c.produit_id = p.id
        """)
        
        columns = [description[0] for description in cursor.description]
        data = cursor.fetchall()
        
        return pd.DataFrame(data, columns=columns)


class AlertSystem:
    """Alert system for urgent orders and thresholds."""
    
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.alerts = []
    
    def check_alerts(self):
        """Check for all alert conditions."""
        self.alerts = []
        
        self._check_urgent_orders()
        self._check_high_quantity_orders()
        self._check_pending_orders_age()
        self._check_suspicious_orders()
        
        return self.alerts
    
    def _check_urgent_orders(self):
        """Check for orders marked as urgent."""
        cursor = self.db.connection.cursor()
        cursor.execute("""
            SELECT c.*, cl.nom as client_nom
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE (c.email_subject LIKE '%urgent%' OR c.informations_supplementaires LIKE '%urgent%')
            AND c.statut = 'en_attente'
        """)
        
        for row in cursor.fetchall():
            self.alerts.append({
                "type": "urgent",
                "severity": "high",
                "message": f"Commande urgente en attente: {row['numero_commande'] or 'N/A'} - {row['client_nom']}",
                "order_id": row["id"]
            })
    
    def _check_high_quantity_orders(self, threshold=10000):
        """Check for unusually high quantity orders."""
        cursor = self.db.connection.cursor()
        cursor.execute("""
            SELECT c.*, cl.nom as client_nom
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.quantite > ? AND c.statut = 'en_attente'
        """, (threshold,))
        
        for row in cursor.fetchall():
            self.alerts.append({
                "type": "high_quantity",
                "severity": "medium",
                "message": f"Commande avec grande quantité ({row['quantite']}): {row['client_nom']}",
                "order_id": row["id"]
            })
    
    def _check_pending_orders_age(self, max_hours=24):
        """Check for orders pending too long."""
        cursor = self.db.connection.cursor()
        cursor.execute("""
            SELECT c.*, cl.nom as client_nom,
                   (julianday('now') - julianday(c.created_at)) * 24 as hours_pending
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.statut = 'en_attente'
            AND (julianday('now') - julianday(c.created_at)) * 24 > ?
        """, (max_hours,))
        
        for row in cursor.fetchall():
            self.alerts.append({
                "type": "pending_too_long",
                "severity": "medium",
                "message": f"Commande en attente depuis {int(row['hours_pending'])}h: {row['client_nom']}",
                "order_id": row["id"]
            })
    
    def _check_suspicious_orders(self):
        """Check for potentially suspicious orders (fraud detection)."""
        cursor = self.db.connection.cursor()
        
        # New clients with very high orders
        cursor.execute("""
            SELECT c.*, cl.nom as client_nom,
                   (SELECT COUNT(*) FROM commandes WHERE client_id = c.client_id) as client_order_count
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.statut = 'en_attente'
            AND c.quantite > 5000
            AND (SELECT COUNT(*) FROM commandes WHERE client_id = c.client_id) <= 1
        """)
        
        for row in cursor.fetchall():
            self.alerts.append({
                "type": "suspicious",
                "severity": "high",
                "message": f"⚠️ Nouveau client avec grande commande: {row['client_nom']} ({row['quantite']} unités)",
                "order_id": row["id"]
            })
        
        # Very low confidence orders
        cursor.execute("""
            SELECT c.*, cl.nom as client_nom
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.confiance < 60 AND c.statut = 'en_attente'
        """)
        
        for row in cursor.fetchall():
            self.alerts.append({
                "type": "low_confidence",
                "severity": "medium",
                "message": f"Commande avec faible confiance ({row['confiance']}%): {row['client_nom']}",
                "order_id": row["id"]
            })


class ReportGenerator:
    """Generate PDF and Excel reports."""
    
    def __init__(self, db: DatabaseManager):
        self.db = db
    
    def export_to_excel_sage(self, filepath="exports/commandes_sage.xlsx", filters=None):
        """Export orders to Excel file in SAGE X3 compatible format."""
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        cursor = self.db.connection.cursor()
        
        # SAGE X3 compatible query with all fields
        query = """
            SELECT 
                c.numero_commande as "Numéro commande",
                c.ligne_commande as "Ligne commande",
                c.site_vente as "Site de vente",
                c.code_client as "Code client",
                cl.nom as "Raison sociale",
                c.code_article as "Code article",
                c.nature_produit as "Désignation",
                c.date_commande as "Date commande",
                c.date_livraison as "Date livraison prévue",
                c.quantite as "Qtée commandée US",
                c.quantite_livree as "Qté Livrée US",
                c.reste_a_livrer as "Reste à livrer US",
                c.quantite_facturee as "Qtée facturée US",
                c.commercial as "Commercial",
                c.type_sac as "Type SAC",
                c.format_sac as "Format LAR.PRE.LON",
                c.type_papier as "Type de papier",
                c.grammage as "Grammage",
                c.laize as "Laize",
                c.impression_client as "Impression client",
                c.prix_unitaire as "Prix unitaire",
                c.prix_total as "Prix total",
                c.devise as "Devise",
                c.statut as "Statut",
                c.confiance as "Confiance IA (%)",
                c.source as "Source",
                c.created_at as "Date création"
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN produits p ON c.produit_id = p.id
        """
        
        if filters:
            conditions = []
            params = []
            if filters.get("status"):
                conditions.append("c.statut = ?")
                params.append(filters["status"])
            if filters.get("client_id"):
                conditions.append("c.client_id = ?")
                params.append(filters["client_id"])
            if filters.get("date_from"):
                conditions.append("c.created_at >= ?")
                params.append(filters["date_from"])
            if filters.get("date_to"):
                conditions.append("c.created_at <= ?")
                params.append(filters["date_to"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            
            query += " ORDER BY c.created_at DESC"
            cursor.execute(query, params)
        else:
            query += " ORDER BY c.created_at DESC"
            cursor.execute(query)
        
        columns = [description[0] for description in cursor.description]
        data = cursor.fetchall()
        
        df = pd.DataFrame(data, columns=columns)
        
        # Create Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Commandes SAGE X3")
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets["Commandes SAGE X3"]
            
            # Apply header formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
    
    def export_to_excel(self, filepath="exports/commandes.xlsx", filters=None):
        """Export orders to Excel file (standard format)."""
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        cursor = self.db.connection.cursor()
        
        query = """
            SELECT 
                c.id,
                c.numero_commande,
                c.ligne_commande,
                c.site_vente,
                c.code_client,
                cl.nom as client,
                c.code_article,
                p.type as produit,
                c.nature_produit,
                c.quantite,
                c.unite,
                c.quantite_livree,
                c.reste_a_livrer,
                c.quantite_facturee,
                c.prix_unitaire,
                c.prix_total,
                c.devise,
                c.date_commande,
                c.date_livraison,
                c.commercial,
                c.type_sac,
                c.format_sac,
                c.type_papier,
                c.grammage,
                c.laize,
                c.impression_client,
                c.statut,
                c.confiance,
                c.source,
                c.created_at
            FROM commandes c
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN produits p ON c.produit_id = p.id
        """
        
        if filters:
            conditions = []
            params = []
            if filters.get("status"):
                conditions.append("c.statut = ?")
                params.append(filters["status"])
            if filters.get("client_id"):
                conditions.append("c.client_id = ?")
                params.append(filters["client_id"])
            if filters.get("date_from"):
                conditions.append("c.created_at >= ?")
                params.append(filters["date_from"])
            if filters.get("date_to"):
                conditions.append("c.created_at <= ?")
                params.append(filters["date_to"])
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        
        columns = [description[0] for description in cursor.description]
        data = cursor.fetchall()
        
        df = pd.DataFrame(data, columns=columns)
        df.to_excel(filepath, index=False, sheet_name="Commandes")
        
        return filepath
    
    def export_to_csv(self, filepath="exports/commandes.csv", filters=None):
        """Export orders to CSV file."""
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Use same query as Excel
        excel_path = filepath.replace('.csv', '_temp.xlsx')
        self.export_to_excel(excel_path, filters)
        
        df = pd.read_excel(excel_path)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        # Clean up temp file
        os.remove(excel_path)
        
        return filepath
    
    def generate_pdf_report(self, filepath="exports/rapport.pdf", period="month"):
        """Generate PDF report with statistics."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import cm
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30
        )
        elements.append(Paragraph("Rapport des Commandes", title_style))
        elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Get stats
        analytics = Analytics(self.db)
        stats = analytics.get_dashboard_stats()
        
        # Summary table
        summary_data = [
            ["Métrique", "Valeur"],
            ["Total Commandes", str(stats["total_orders"])],
            ["Commandes Validées", str(stats["validated_orders"])],
            ["Commandes En Attente", str(stats["pending_orders"])],
            ["Commandes Rejetées", str(stats["rejected_orders"])],
            ["Total Clients", str(stats["total_clients"])],
            ["Chiffre d'Affaires Total", f"{stats['total_revenue']:.2f} MAD"],
            ["Valeur Moyenne Commande", f"{stats['avg_order_value']:.2f} MAD"],
        ]
        
        table = Table(summary_data, colWidths=[8*cm, 6*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 30))
        
        # Top clients
        elements.append(Paragraph("Top Clients", styles['Heading2']))
        if stats["top_clients"]:
            client_data = [["Client", "Commandes", "Total Dépensé"]]
            for client in stats["top_clients"][:5]:
                client_data.append([
                    client["nom"] or "N/A",
                    str(client["order_count"]),
                    f"{client['total_spent'] or 0:.2f} MAD"
                ])
            
            client_table = Table(client_data, colWidths=[7*cm, 4*cm, 4*cm])
            client_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#764ba2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
            ]))
            elements.append(client_table)
        
        doc.build(elements)
        return filepath


class AIPredictor:
    """AI-powered predictions and analysis."""
    
    def __init__(self, db: DatabaseManager):
        self.db = db
        self.client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    
    def predict_client_behavior(self, client_name):
        """Predict client's future ordering behavior."""
        history = ClientHistory(self.db)
        preferences = history.get_client_preferences(client_name)
        
        if not preferences:
            return None
        
        prediction = {
            "client": client_name,
            "predicted_next_order_days": None,
            "predicted_quantity": preferences["average_quantity"],
            "likely_products": list(preferences["favorite_products"].keys())[:3],
            "client_value": "high" if preferences["total_spent"] > 10000 else "medium" if preferences["total_spent"] > 1000 else "low"
        }
        
        if preferences["order_frequency_days"]:
            prediction["predicted_next_order_days"] = int(preferences["order_frequency_days"])
        
        return prediction
    
    def classify_email_priority(self, email_subject, email_body):
        """Classify email priority using AI."""
        prompt = f"""Analyse cet email et détermine sa priorité.

Sujet: {email_subject}
Contenu: {email_body[:500]}

Réponds en JSON avec:
- priority: "high", "medium", ou "low"
- is_purchase_order: true/false
- is_urgent: true/false
- confidence: 0-100

JSON uniquement:"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=200
            )
            
            result = response.choices[0].message.content.strip()
            if result.startswith("```"):
                result = result.split("```")[1].replace("json", "").strip()
            
            return json.loads(result)
        except:
            return {"priority": "medium", "is_purchase_order": False, "is_urgent": False, "confidence": 50}
    
    def detect_anomalies(self, order_data):
        """Detect anomalies in order data."""
        anomalies = []
        
        # Check quantity anomaly
        cursor = self.db.connection.cursor()
        cursor.execute("SELECT AVG(quantite), MAX(quantite) FROM commandes WHERE quantite IS NOT NULL")
        avg_qty, max_qty = cursor.fetchone()
        
        if order_data.get("quantite") and avg_qty:
            if order_data["quantite"] > avg_qty * 3:
                anomalies.append({
                    "type": "quantity_anomaly",
                    "message": f"Quantité ({order_data['quantite']}) très supérieure à la moyenne ({avg_qty:.0f})"
                })
        
        # Check price anomaly
        if order_data.get("prix_total") and order_data.get("quantite"):
            unit_price = order_data["prix_total"] / order_data["quantite"]
            cursor.execute("""
                SELECT AVG(prix_total / quantite) 
                FROM commandes 
                WHERE prix_total IS NOT NULL AND quantite > 0
            """)
            avg_unit = cursor.fetchone()[0]
            
            if avg_unit and (unit_price < avg_unit * 0.3 or unit_price > avg_unit * 3):
                anomalies.append({
                    "type": "price_anomaly",
                    "message": f"Prix unitaire ({unit_price:.2f}) inhabituel par rapport à la moyenne ({avg_unit:.2f})"
                })
        
        return anomalies


def test_analytics():
    """Test analytics features."""
    print("=" * 50)
    print("🧪 Test des fonctionnalités analytiques")
    print("=" * 50)
    
    db = DatabaseManager()
    db.connect()
    db.init_database()
    
    # Test stats
    print("\n📊 Statistiques du dashboard:")
    analytics = Analytics(db)
    stats = analytics.get_dashboard_stats()
    print(f"   Total commandes: {stats['total_orders']}")
    print(f"   En attente: {stats['pending_orders']}")
    print(f"   Validées: {stats['validated_orders']}")
    
    # Test alerts
    print("\n🚨 Alertes:")
    alert_system = AlertSystem(db)
    alerts = alert_system.check_alerts()
    for alert in alerts:
        print(f"   [{alert['severity'].upper()}] {alert['message']}")
    
    if not alerts:
        print("   Aucune alerte")
    
    # Test export
    print("\n📁 Export Excel:")
    reporter = ReportGenerator(db)
    excel_path = reporter.export_to_excel()
    print(f"   ✅ {excel_path}")
    
    db.disconnect()
    print("\n✅ Test terminé")


if __name__ == "__main__":
    test_analytics()
